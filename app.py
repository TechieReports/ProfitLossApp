import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

def process_spend_and_revenue(spend_files, revenue_files):
    spends_df = pd.concat([pd.read_excel(file) for file in spend_files], ignore_index=True)
    revenue_raw_df = pd.concat([pd.read_csv(file) for file in revenue_files], ignore_index=True)

    # Aggregate raw revenue data
    revenue_agg = revenue_raw_df.groupby(['campid', 'date'], as_index=False).agg({
        'clicks': 'sum',
        'estimated_earnings': 'sum'
    })
    revenue_agg.rename(columns={
        'campid': 'Camp ID',
        'date': 'Date',
        'clicks': 'Total Clicks',
        'estimated_earnings': 'Revenue'
    }, inplace=True)
    revenue_agg['Date'] = pd.to_datetime(revenue_agg['Date'])

    # Extract campaign IDs from spend file
    spends_df['Camp ID'] = spends_df['Ad set name'].str.extract(r'\((\d+)\)').astype(int)
    spends_df['Date'] = pd.to_datetime(spends_df['Day'])

    # Handle "Cost per result" or "Cost per purchase" and rename to "CPR"
    if 'Cost per result' in spends_df.columns:
        spends_df.rename(columns={'Cost per result': 'CPR'}, inplace=True)
    elif 'Cost per purchase' in spends_df.columns:
        spends_df.rename(columns={'Cost per purchase': 'CPR'}, inplace=True)
    else:
        spends_df['CPR'] = 0  # Default CPR value if both columns are missing

    # Merge spend and revenue data
    merged_data = spends_df.merge(revenue_agg, how='left', on=['Camp ID', 'Date'])
    merged_data['RPC'] = merged_data['Revenue'] / merged_data['Total Clicks']
    merged_data['RPC'] = merged_data['RPC'].replace([float('inf'), -float('inf')], 0).fillna(0)
    merged_data['Profit/Loss'] = merged_data['Revenue'] - merged_data['Amount spent (USD)']

    # Sort by Date (oldest to newest) and Camp ID (smallest to largest)
    merged_data.sort_values(by=['Date', 'Camp ID'], inplace=True)

    # Reorder columns
    column_order = ['Camp ID', 'Ad set name', 'Date', 'Amount spent (USD)', 'Revenue', 'CPR', 'RPC', 'Profit/Loss']
    merged_data = merged_data[column_order]

    return merged_data

def create_excel_file(data):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Profit Loss Analysis"

    # Add headers with formatting
    headers = ['Camp ID', 'Campaign Name', 'Date', 'Spend', 'Revenue', 'CPR', 'RPC', 'Profit/Loss']
    header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    header_font = Font(bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Freeze header row
    ws.freeze_panes = ws['A2']

    # Add data rows and format the date column
    for row_idx, row in enumerate(dataframe_to_rows(data, index=False, header=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # Format Date column to "MMM DD"
            if col_idx == 3:  # Date is the 3rd column
                cell.number_format = "MMM DD"

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value is not None)
        adjusted_width = max_length + 2
        ws.column_dimensions[col[0].column_letter].width = adjusted_width

    # Conditional formatting for Profit/Loss
    profit_loss_col_index = headers.index('Profit/Loss') + 1
    for row in ws.iter_rows(min_row=2, min_col=profit_loss_col_index, max_col=profit_loss_col_index):
        for cell in row:
            if cell.value is not None:
                if cell.value > 0:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif cell.value < 0:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    wb.save(output)
    output.seek(0)
    return output

# Streamlit UI
st.title("Profit/Loss Analysis Tool with Multi-File Upload")

spend_files = st.file_uploader("Upload Spend Files (Excel)", type=["xlsx"], accept_multiple_files=True)
revenue_files = st.file_uploader("Upload Revenue Files (CSV)", type=["csv"], accept_multiple_files=True)

if spend_files and revenue_files:
    data = process_spend_and_revenue(spend_files, revenue_files)

    # Filters
    st.write("Filters:")
    date_range = st.date_input("Date Range", [])
    available_campaigns = data['Camp ID'].unique().tolist()
    selected_campaigns = st.multiselect("Select Campaigns", options=available_campaigns, default=available_campaigns)
    custom_campaigns = st.text_input("Or Enter Campaign IDs (comma-separated):")

    # Parse custom Campaign IDs
    if custom_campaigns:
        custom_campaigns = [int(c.strip()) for c in custom_campaigns.split(",") if c.strip().isdigit()]
        selected_campaigns = selected_campaigns + custom_campaigns

    # Apply Filters Dynamically
    filtered_data = data.copy()
    if date_range:
        filtered_data = filtered_data[
            (filtered_data['Date'] >= pd.to_datetime(date_range[0])) &
            (filtered_data['Date'] <= pd.to_datetime(date_range[1]))
        ]
    if selected_campaigns:
        filtered_data = filtered_data[filtered_data['Camp ID'].isin(selected_campaigns)]

    st.write("Filtered Data:")
    st.dataframe(filtered_data)

    if st.button("Download Filtered Data"):
        output_file = create_excel_file(filtered_data)
        st.download_button(
            label="Download Profit/Loss Report",
            data=output_file,
            file_name="Profit_Loss_Analysis.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
